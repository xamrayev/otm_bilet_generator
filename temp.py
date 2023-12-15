import openpyxl

def savollar_from_xls(file=str):
    '''
    Exel faylidan ma'lumotlarni olish va ular asosida biletlar tuzish uchun ishlatiladi.
    funksiya file = file_to_path metodi da ishlaydi
    file_to_path o'rniga savollar joylashgan xls fayliga yo'l ko'rsating
    '''
    filename = file
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    first_column_data = []
    for row in sheet.iter_rows():
        first_column_data.append(row[0].value)
    if sheet.max_row > 1:
        first_column_data = first_column_data[1:]
    return first_column_data

# print(len(savollar_from_xls(file="savollar.xlsx")))

savollar = [f"savol_{x}" for x in range(25)]
biletlar = int(input())
j=0
bilet=0
for i in range(biletlar):
    bilet+=1
    print("\n",bilet,"\n")
    print(f"1) {savollar[j+0]}")
    print(f"2) {savollar[j+1]}")
    print(f"3) {savollar[j+2]}")
    print(f"4) {savollar[j+3]}")
    print(f"5) {savollar[j+4]}")
    j+=5
    
