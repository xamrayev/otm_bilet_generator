from test_2_list import tayyorlash
from doc import test_to_doc
from savol_2_doc import biletlar_ru, savollar_from_xls
from openpyxl import Workbook

import tkinter as tk
import os


from tkinter import filedialog, messagebox, ttk 
import customtkinter as ctk

ctk.set_appearance_mode("System")        
 
ctk.set_default_color_theme("green") 

def open_file_path():
    file_path = filedialog.askopenfilename()
    if file_path:
        file_path_entry.delete(0, tk.END)
        file_path_entry.insert(0, file_path)


def open_file_path2():
    file_path = filedialog.askopenfilename()
    if file_path:
        file_path_entry2.delete(0, tk.END)
        file_path_entry2.insert(0, file_path)

def tayyorla_clicked():
    try:
        # Gather values frcom CTkentry widgets
        kerakli_variantlar_soni = eval(variant_soni_entry.get())
        variantda_test_soni = eval(variant_test_soni_entry.get())
        til = til_var.get()
        kafedra_nomi = kafedra_nomi_entry.get()
        fan = fan_nomi_entry.get()
        guruh = guruh_nomi_entry.get()
        semestr = semestr_entry.get()
        tuzuvchi = tuzuvchi_entry.get()
        zav_kaf = zav_kaf_entry.get()
        fayl = file_path_entry.get()

        testlar = tayyorlash(test_fayl=fayl, kerakli_variantlar_soni=kerakli_variantlar_soni,variantda_test_soni=variantda_test_soni)
        test_to_doc(
            variantlar=testlar,
            til = til,
            kafedra_nomi=kafedra_nomi,
            guruh = guruh,
            fan = fan,
            semestr = semestr,
            tuzuvchi = tuzuvchi,
            zav_kaf = zav_kaf,
        )
        current_directory = os.getcwd()

        messagebox.showinfo("Test yaratildi", f"Test - {current_directory} papkasida {fan} nomi bilan alohida papkaga saqlandi")
    except:
        messagebox.showwarning("Xatolik", f"Dastur ishlashi uchun barcha bo'limlarni to'ldiring")


def tayyorla2_clicked():
    try:
        til = til_var2.get()
        kafedra_nomi = kafedra_nomi_entry2.get()
        fanimiz = fan_nomi_entry2.get()
        semestr = semestr_entry2.get()
        tuzuvchi = tuzuvchi_entry2.get()
        kafedra_mudiri = zav_kaf_entry2.get()
        bilet_soni = eval(variant_soni_entry2.get())
        savollar_soni = eval(variant_test_soni_entry2.get())
        current_directory = os.getcwd()
        savol_file = file_path_entry2.get()
        savollarim = savollar_from_xls(savol_file)
        
        biletlar_ru(
            til=til,
            bilet_soni=bilet_soni, 
            savollar=savollarim, 
            fan=fanimiz,
            semestr=semestr, 
            kafedra=kafedra_nomi, 
            tuzuvchi=tuzuvchi, 
            zav_kaf=kafedra_mudiri, 
            savollar_soni=savollar_soni
            
        )
        messagebox.showinfo("Test yaratildi", f"Savollar {current_directory} papkasida {fanimiz}_biletlar.docx nomi bilan saqlandi")
    except:

        messagebox.showwarning("Xatolik", f"Dastur ishlashi uchun barcha bo'limlarni to'ldiring")

def test_template():
    wb = Workbook()

    # Select the active worksheet
    ws = wb.active

    # Set the column headers
    ws['A1'] = 'Savollar'
    ws['B1'] = 'a'
    ws['C1'] = 'b'
    ws['D1'] = 'c'
    ws['E1'] = 'd'
    current_directory = os.getcwd()
    wb.save(f'{current_directory}/test_example.xlsx')
    messagebox.showinfo("Namuna yaratildi", f"Namuna - {current_directory} papkasida test_example.xlsx nomi bilan saqlandi")


def savol_template():
    wb = Workbook()

    # Select the active worksheet
    ws = wb.active

    # Set the column headers
    ws['A1'] = 'No'
    ws['B1'] = 'Savol'
    
    current_directory = os.getcwd()
    wb.save(f'{current_directory}/savol_example.xlsx')
    messagebox.showinfo("Namuna yaratildi", f"Namuna - {current_directory} papkasida savol_example.xlsx nomi bilan saqlandi")


root = ctk.CTk()
root.title("Assistent 1.0")

notebook = ttk.Notebook()
notebook.pack(expand=True, fill="both")

frame1 = ttk.Frame(notebook)
frame2 = ttk.Frame(notebook)
frame1.pack(fill="both", expand=True)
frame2.pack(fill="both", expand=True)
notebook.add(frame1, text="Test", compound="center")
notebook.add(frame2, text="Savollar", compound="center")

# Variant Soni
variant_soni_label = ctk.CTkLabel(frame1, text="Variant Soni:")
variant_soni_label.grid(row=5, column=0,)
variant_soni_entry = ctk.CTkEntry(frame1)
variant_soni_entry.grid(row=5, column=1,)

# Variant Test Soni
variant_test_soni_label = ctk.CTkLabel(frame1, text="Testlar Soni:")
variant_test_soni_label.grid(row=6, column=0,)
variant_test_soni_entry = ctk.CTkEntry(frame1)
variant_test_soni_entry.grid(row=6, column=1,)

# Til
til_label = ctk.CTkLabel(frame1, text="Til:")
til_label.grid(row=0, column=0, padx=5)
til_var = tk.StringVar()
til_options = ["uz", "ru"]
for index, til in enumerate(til_options):
    ctk.CTkRadioButton(frame1, text=til, variable=til_var, value=til, width=5).grid(row=0, column=index + 1,)

# Kafedra nomi
kafedra_nomi_label = ctk.CTkLabel(frame1, text="Kafedra nomi:")
kafedra_nomi_label.grid(row=1, column=0,  sticky="W")
kafedra_nomi_entry = ctk.CTkEntry(frame1)
kafedra_nomi_entry.grid(row=1, column=1,)


# Fan nomi
fan_nomi_label = ctk.CTkLabel(frame1, text="Fan nomi:")  # Added Fan nomi
fan_nomi_label.grid(row=2, column=0,  sticky="W")  # Added Fan nomi
fan_nomi_entry = ctk.CTkEntry(frame1)  # Added Fan nomi
fan_nomi_entry.grid(row=2, column=1,)  # Added Fan nomi

# Guruh nomi
guruh_nomi_label = ctk.CTkLabel(frame1, text="Guruh:")
guruh_nomi_label.grid(row=3, column=0,  sticky="W")
guruh_nomi_entry = ctk.CTkEntry(frame1)
guruh_nomi_entry.grid(row=3, column=1, )

# Semestr
semestr_label = ctk.CTkLabel(frame1, text="Semestr:")
semestr_label.grid(row=4, column=0,sticky="W")
semestr_entry = ctk.CTkEntry(frame1)
semestr_entry.grid(row=4, column=1,)

# Tuzuvchi
tuzuvchi_label = ctk.CTkLabel(frame1, text="Tuzuvchi:")
tuzuvchi_label.grid(row=7, column=0, sticky="W")
tuzuvchi_entry = ctk.CTkEntry(frame1)
tuzuvchi_entry.grid(row=7, column=1,)

# Zavod Kafedra
zav_kaf_label = ctk.CTkLabel(frame1, text="Kafedra mudiri:")
zav_kaf_label.grid(row=8, column=0, sticky="W")
zav_kaf_entry = ctk.CTkEntry(frame1)
zav_kaf_entry.grid(row=8, column=1,)

# CTkButton to open file dialog for file path
open_file_button = ctk.CTkButton(frame1, text="Open File", command=open_file_path, width=10)
open_file_button.grid(row=9, column=0, pady=10, sticky="W")

# CTkEntry for file path
file_path_entry = ctk.CTkEntry(frame1)
file_path_entry.grid(row=9, column=1, sticky="W")

# "Tayyorlca" CTkbutton
tayyorla_button = ctk.CTkButton(frame1, text="Tayyorla", command=tayyorla_clicked, width=10)
tayyorla_button.grid(row=10, column=0,)
namuna_test_button = ctk.CTkButton(frame1, text="Namuna fayl", command=test_template, width=10)
namuna_test_button.grid(row=10, column=1, sticky="W")


# Til
til_label2 = ctk.CTkLabel(frame2, text="Til:")
til_label2.grid(row=0, column=0, padx=5)
til_var2 = tk.StringVar()
til_options = ["uz", "ru"]
for index, til in enumerate(til_options):
    ctk.CTkRadioButton(frame2, text=til, variable=til_var2, value=til, width=5).grid(row=0, column=index + 1, )

# Kafedra nomi
kafedra_nomi_label2 = ctk.CTkLabel(frame2, text="Kafedra nomi:")
kafedra_nomi_label2.grid(row=1, column=0,  sticky="W")
kafedra_nomi_entry2 = ctk.CTkEntry(frame2)
kafedra_nomi_entry2.grid(row=1, column=1,)

# Fan nomi
fan_nomi_label2 = ctk.CTkLabel(frame2, text="Fan nomi:")  # Added Fan nomi
fan_nomi_label2.grid(row=2, column=0,  sticky="W")  # Added Fan nomi
fan_nomi_entry2 = ctk.CTkEntry(frame2)  # Added Fan nomi
fan_nomi_entry2.grid(row=2, column=1,)  # Added Fan nomi

# Guruh nomi
guruh_nomi_label2 = ctk.CTkLabel(frame2, text="Guruh:")
guruh_nomi_label2.grid(row=3, column=0,  sticky="W")
guruh_nomi_entry2 = ctk.CTkEntry(frame2)
guruh_nomi_entry2.grid(row=3, column=1, )



# Semestr
semestr_label2 = ctk.CTkLabel(frame2, text="Semestr:")
semestr_label2.grid(row=4, column=0,sticky="W")
semestr_entry2 = ctk.CTkEntry(frame2)
semestr_entry2.grid(row=4, column=1,)


# Variant Soni
variant_soni_label2 = ctk.CTkLabel(frame2, text="Variant Soni:")
variant_soni_label2.grid(row=5, column=0,)
variant_soni_entry2 = ctk.CTkEntry(frame2)
variant_soni_entry2.grid(row=5, column=1,)

# Variant Test Soni
variant_test_soni_label2 = ctk.CTkLabel(frame2, text="Savollar Soni:")
variant_test_soni_label2.grid(row=6, column=0,)
variant_test_soni_entry2 = ctk.CTkEntry(frame2)
variant_test_soni_entry2.grid(row=6, column=1,)

# Tuzuvchi
tuzuvchi_label2 = ctk.CTkLabel(frame2, text="Tuzuvchi:")
tuzuvchi_label2.grid(row=7, column=0, sticky="W")
tuzuvchi_entry2 = ctk.CTkEntry(frame2)
tuzuvchi_entry2.grid(row=7, column=1,)

# Zavod Kafedra
zav_kaf_label2 = ctk.CTkLabel(frame2, text="Kafedra mudiri:")
zav_kaf_label2.grid(row=8, column=0, sticky="W")
zav_kaf_entry2 = ctk.CTkEntry(frame2)
zav_kaf_entry2.grid(row=8, column=1,)

# CTkButton to open file dialog for file path
open_file_button = ctk.CTkButton(frame2, text="Open File", command=open_file_path2, width=10)
open_file_button.grid(row=9, column=0, pady=10)

# CTkEntry for file path
file_path_entry2 = ctk.CTkEntry(frame2)
file_path_entry2.grid(row=9, column=1,)

# "Tayyorlca" CTkbutton
tayyorla_button2 = ctk.CTkButton(frame2, text="Tayyorla", command=tayyorla2_clicked, width=10)
tayyorla_button2.grid(row=10, column=0,sticky="W")

namuna_savol_button = ctk.CTkButton(frame2, text="Namuna fayl", command=savol_template, width=10)
namuna_savol_button.grid(row=10, column=1,sticky="W")


root.mainloop()
